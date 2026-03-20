#!/usr/bin/env python3
"""
Import OEM Catalyst Database (Bosal) from Excel into JSON for the platform.

Supports V3, V4, and future workbooks with the same sheet names. V4 adds "Source Traceability".

Usage:
  python3 scripts/import-oem-catalyst-database.py [path/to/OEM_Catalyst_Database_V4_Bosal.xlsx]

Default input: ~/Downloads/OEM_Catalyst_Database_V4_Bosal.xlsx
Output: src/lib/catsizer/oem-database/data/*.json
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "src" / "lib" / "catsizer" / "oem-database" / "data"

# ECS Component Database — column keys (row index 3 in sheet, 0-based row 3)
ECS_KEYS = [
    "oemGroup",
    "brand",
    "engineFamily",
    "engineCodes",
    "fuel",
    "displacementL",
    "cylinders",
    "powerKw",
    "emissionStandard",
    "years",
    "vehicleExamples",
    "productionVolumeEu",
    "componentNumber",
    "componentType",
    "position",
    "substrate",
    "substrateSupplier",
    "diameterMm",
    "lengthMm",
    "volumeL",
    "cpsi",
    "wallMil",
    "geometricSaM2PerL",
    "ofaPercent",
    "wcLayers",
    "wcTotalGPerL",
    "l1Support",
    "l1SaM2PerG",
    "l1Stabiliser",
    "l1Promoter",
    "l1OscMaterial",
    "l1OscComposition",
    "l1OscGPerL",
    "l1Pgm",
    "l1Extra",
    "l1WcGPerL",
    "l2Support",
    "l2SaM2PerG",
    "l2Stabiliser",
    "l2Promoter",
    "l2OscMaterial",
    "l2OscComposition",
    "l2OscGPerL",
    "l2Pgm",
    "l2Extra",
    "l2WcGPerL",
    "totalOscGPerL",
    "ptGPerL",
    "pdGPerL",
    "rhGPerL",
    "totalPgmGPerL",
    "ptGPerBrick",
    "pdGPerBrick",
    "rhGPerBrick",
    "t50CoC",
    "t50HcC",
    "t50NoxC",
    "maxExoC",
    "bpAtRatedKpa",
    "agingProtocol",
    "trend",
    "confidence",
    "source",
]


def cell_to_json(v):
    if v is None:
        return None
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 6)
    return str(v).strip() if isinstance(v, str) else v


def row_to_obj(keys, row):
    out = {}
    for i, k in enumerate(keys):
        val = row[i] if i < len(row) else None
        out[k] = cell_to_json(val)
    return out


def parse_ecs_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return []
    header_len = len(ECS_KEYS)
    records = []
    current_section = None
    for idx, row in enumerate(rows):
        if idx < 4:
            continue
        r = list(row)[:header_len]
        non_empty = sum(1 for x in r if x is not None and str(x).strip() != "")
        # Section heading row: only first cell meaningful, or ▸ prefix
        first = r[0]
        first_s = str(first).strip() if first is not None else ""
        is_section = non_empty <= 2 and first_s and (
            first_s.startswith("▸") or (r[1] is None and r[12] is None and non_empty <= 2)
        )
        if is_section and first_s:
            current_section = first_s.replace("▸", "").strip()
            continue
        # Data row: must have component # or brand + engine
        if r[12] is None and r[1] is None:
            continue
        if r[1] is None and r[2] is None:
            continue
        obj = row_to_obj(ECS_KEYS, r)
        if current_section:
            obj["sheetSection"] = current_section
        records.append(obj)
    return records


def parse_generic_sheet(ws, header_row_index: int):
    rows = list(ws.iter_rows(values_only=True))
    if header_row_index >= len(rows):
        return []
    headers = [cell_to_json(c) for c in rows[header_row_index]]
    while headers and headers[-1] is None:
        headers.pop()
    keys = []
    used = {}
    for h in headers:
        if h is None:
            keys.append(f"col_{len(keys)}")
            continue
        base = "".join(c if c.isalnum() else "_" for c in str(h))[:60].strip("_").lower()
        if not base:
            base = f"col_{len(keys)}"
        k = base
        if k in used:
            used[k] += 1
            k = f"{k}_{used[k]}"
        else:
            used[k] = 0
        keys.append(k)
    out = []
    for row in rows[header_row_index + 1 :]:
        r = list(row)[: len(keys)]
        if all(x is None or str(x).strip() == "" for x in r):
            continue
        obj = row_to_obj(keys, r + [None] * max(0, len(keys) - len(r)))
        if any(v is not None for v in obj.values()):
            out.append(obj)
    return out


def infer_version(filename: str) -> str:
    upper = filename.upper()
    for tag in ("V6", "V5", "V4", "V3", "V2", "V1"):
        if tag in upper:
            return tag
    return "unknown"


def main():
    default_xlsx = Path.home() / "Downloads" / "OEM_Catalyst_Database_V4_Bosal.xlsx"
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx
    if not src.is_file():
        print(f"Excel not found: {src}", file=sys.stderr)
        sys.exit(1)

    version = infer_version(src.name)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src, read_only=True, data_only=True)

    ecs = parse_ecs_sheet(wb["ECS Component Database"])
    washcoat = parse_generic_sheet(wb["Washcoat Chemistry Detail"], 2)
    am_guidance = parse_generic_sheet(wb["AM Design Guidance"], 2)
    pt_pd = parse_generic_sheet(wb["Pt-Pd Substitution"], 2)
    arch = parse_generic_sheet(wb["System Architecture Map"], 2)

    source_traceability = []
    if "Source Traceability" in wb.sheetnames:
        source_traceability = parse_generic_sheet(wb["Source Traceability"], 3)

    manifest = {
        "databaseVersion": version,
        "sourceFile": src.name,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "counts": {
            "ecsComponents": len(ecs),
            "washcoatChemistry": len(washcoat),
            "amDesignGuidance": len(am_guidance),
            "ptPdSubstitution": len(pt_pd),
            "systemArchitecture": len(arch),
            "sourceTraceability": len(source_traceability),
        },
        "description": "European OEM ECS component-level reference for Bosal AM homologation copilot",
    }

    (OUT_DIR / "manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "ecs-components.json").write_text(
        json.dumps(ecs, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "washcoat-chemistry.json").write_text(
        json.dumps(washcoat, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "am-design-guidance.json").write_text(
        json.dumps(am_guidance, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "pt-pd-substitution.json").write_text(
        json.dumps(pt_pd, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "system-architecture.json").write_text(
        json.dumps(arch, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "source-traceability.json").write_text(
        json.dumps(source_traceability, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    wb.close()
    print(
        f"Wrote {OUT_DIR} — {version} ECS: {len(ecs)}, source traceability: {len(source_traceability)}"
    )


if __name__ == "__main__":
    main()
